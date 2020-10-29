import numpy
import pandas
import openpyxl
from neo4j import GraphDatabase


def loadDataSet():  #加载数据集
    wb = openpyxl.load_workbook('cause.xlsx')
    # 读取工作表
    ws = wb['Sheet1']
    Cause = []
    Effect = []

    for col in ws['A']:
        Cause.append(col.value)
        print(col)
    for col in ws['C']:
        Effect.append(col.value)

    print(Cause)
    print(Effect)
    return Cause, Effect



class Example(object):

    def __init__(self, uri, user, password):
        self._driver = GraphDatabase.driver(uri, auth=(user, password))

    def close(self):
        self._driver.close()
    #创建实体节点
    def create_Cause(cls, tx, Name):
        tx.run("CREATE (:Cause {Name: $Name})", Name=Name)

    def create_Effect(cls, tx, Name):
        tx.run("CREATE (:Effect {Name: $Name})", Name = Name)

    #创建关系
    def Cause_Effect(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Cause {Name: $node_name1}) "
               "MATCH (node2: Effect {Name: $node_name2}) "
               "CREATE (node1)-[:cause]->(node2)",
               node_name1=node_name1, node_name2=node_name2)



    #创建Driver
    def create_Cause_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Cause, message)

    def create_Effect_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Effect, message)


    def create_Cause_Effect_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Cause_Effect, message1, message2)


exam=Example("bolt://localhost:7687","neo4j","myneo4j")

Cause, Effect = loadDataSet()


# #导入节点
# for each in Factor:
#     exam.create_Company_driver(each)
# for each in Operating_Status:
#     exam.create_Ticker_driver(each)
# for each in Social_Factors:
#     exam.create_Industry_driver(each)
# for each in Political_Factors:
#     exam.create_Industry_Category_driver(each)
#
# #导入公司和股票的关系
# for i in range(0, len(Company)):
#     exam.create_Company_Ticker_driver(Company[i], Ticker[i])
#
# #导入公司和行业的关系
# exam.create_Company_Industry_driver(Company[0], Primary_Industry[0])
# for i in range(1, len(Company)):
#     if(Company[i] != Company[i-1]):
#         exam.create_Company_Industry_driver(Company[i], Primary_Industry[i])
#
# #行业分类
# s = []
# for i in range(0, len(Primary_Industry)):
#     if(Primary_Industry[i] not in s):
#         exam.create_Industry_Classify_driver(Primary_Industry[i], Industry_Category[i])
#         s.append(Primary_Industry[i])


s1 = []
s2 = []
for i in range(0, len(Cause)):
    if Cause[i] not in s1:
        s1.append(Cause[i])
        exam.create_Cause_driver(Cause[i])
    if Effect[i] not in s2:
        s2.append(Effect[i])
        exam.create_Effect_driver(Effect[i])
    exam.create_Cause_Effect_driver(Cause[i], Effect[i])

