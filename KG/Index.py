import numpy
import pandas
import openpyxl
from neo4j import GraphDatabase


def loadDataSet():  #加载数据集
    wb = openpyxl.load_workbook('index.xlsx')
    # 读取工作表
    ws = wb['Sheet1']
    Factor = []
    Factor.append('Operating Status')
    Factor.append('Social Factors')
    Factor.append('Political Factors')
    Factor.append('Transaction Situation')
    Operating_Status = []
    Social_Factors = []
    Political_Factors= []
    Transaction_Situation = []
    for col in ws['A']:
        Operating_Status.append(col.value)
    Operating_Status.remove('Operating Status')
    for col in ws['B']:
        Social_Factors.append(col.value)
    Social_Factors.remove('Social Factors')
    for col in ws['C']:
        Political_Factors.append(col.value)
    Political_Factors.remove('Political Factors')
    for col in ws['D']:
        Transaction_Situation.append(col.value)
    Transaction_Situation.remove('Transaction Situation')

    print("Factor", Factor)
    print("Operating_Status", Operating_Status)
    print("Social_Factors", Social_Factors)
    print("Political_Factors", Political_Factors)
    print("Transaction_Situation", Transaction_Situation)

    return  Factor, Operating_Status, Social_Factors, Political_Factors, Transaction_Situation



class Example(object):

    def __init__(self, uri, user, password):
        self._driver = GraphDatabase.driver(uri, auth=(user, password))

    def close(self):
        self._driver.close()
    #创建实体节点
    def create_Factor(cls, tx, Name):
        tx.run("CREATE (:Factor {Name: $Name})", Name=Name)

    def create_Operating_Status(cls, tx, Operating_Status):
        tx.run("CREATE (:Operating_Status {Name: $Operating_Status})", Operating_Status = Operating_Status)

    def create_Social_Factors(cls, tx, Social_Factors):
        tx.run("CREATE (:Social_Factors {Name: $Social_Factors})", Social_Factors = Social_Factors)

    def create_Political_Factors(cls, tx, Political_Factors):
        tx.run("CREATE (:Political_Factors {Name: $Political_Factors})", Political_Factors = Political_Factors)

    def create_Transaction_Situation(cls, tx, Transaction_Situation):
        tx.run("CREATE (:Transaction_Situation {Name: $Transaction_Situation})", Transaction_Situation = Transaction_Situation)

    def create_Financial_elements(cls, tx, Name):
        tx.run("CREATE (:Financial_elements {Name: $Name})", Name=Name)


    #创建关系
    def Operating_Status_index(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Factor {Name: $node_name1}) "
               "MATCH (node2: Operating_Status {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Social_Factors_index(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Factor {Name: $node_name1}) "
               "MATCH (node2: Social_Factors {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Political_Factors_index(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Factor {Name: $node_name1}) "
               "MATCH (node2: Political_Factors {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Transaction_Situation_index(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Factor {Name: $node_name1}) "
               "MATCH (node2: Transaction_Situation {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Financial_elements_Factor(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Financial_elements {Name: $node_name1}) "
               "MATCH (node2: Factor {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    #创建节点Driver
    def create_Factor_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Factor, message)

    def create_Operating_Status_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Operating_Status, message)

    def create_Social_Factors_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Social_Factors, message)

    def create_Political_Factors_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Political_Factors, message)

    def create_Transaction_Situation_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Transaction_Situation, message)

    def create_Financial_elements_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Financial_elements, message)



    # 创建关系Driver
    def Operating_Status_index_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Operating_Status_index, message1, message2)

    def Social_Factors_index_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Social_Factors_index, message1, message2)

    def Political_Factors_index_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Political_Factors_index, message1, message2)

    def Transaction_Situation_index_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Transaction_Situation_index, message1, message2)

    def Financial_elements_Factor_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Financial_elements_Factor, message1, message2)

exam=Example("bolt://18.166.174.144:5002","neo4j","chengdu80uestc")

Factor, Operating_Status, Social_Factors, Political_Factors, Transaction_Situation = loadDataSet()


#导入节点
for each in Factor:
    exam.create_Factor_driver(each)
Operating_Status.remove(None)
Operating_Status.remove(None)
print(Operating_Status)
for each in Operating_Status:
    exam.create_Operating_Status_driver(each)
for each in Social_Factors:
    exam.create_Social_Factors_driver(each)
for each in Political_Factors:
    exam.create_Political_Factors_driver(each)
for each in Transaction_Situation:
    exam.create_Transaction_Situation_driver(each)
exam.create_Financial_elements_driver('Financial elements')


for i in range(0, len(Operating_Status)):
    exam.Operating_Status_index_driver(Factor[0], Operating_Status[i])

for i in range(0, len(Social_Factors)):
        exam.Social_Factors_index_driver(Factor[1], Social_Factors[i])

for i in range(0, len(Political_Factors)):
        exam.Political_Factors_index_driver(Factor[2], Political_Factors[i])

for i in range(0, len(Transaction_Situation)):
        exam.Transaction_Situation_index_driver(Factor[3], Transaction_Situation[i])

for item in Factor:
        exam.Financial_elements_Factor_driver('Financial elements', item)



