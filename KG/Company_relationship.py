import numpy
import pandas
import openpyxl
from neo4j import GraphDatabase


def loadDataSet():  #加载数据集
    wb = openpyxl.load_workbook('Company_relationship.xlsx')
    # 读取工作表
    ws = wb['Sheet1']
    Company = []
    Ticker = []
    Primary_Industry = []
    Industry_category = []
    for col in ws['A']:
        Company.append(col.value)
    Company.remove('Company Name')
    for col in ws['B']:
        Ticker.append(col.value)
    Ticker.remove('Ticker')
    for col in ws['C']:
        Primary_Industry.append(col.value)
    Primary_Industry.remove('Primary Industry')
    for col in ws['D']:
        Industry_category.append(col.value)
    Industry_category.remove('Industry category')


    print("Company", Company)
    print("Ticker", Ticker)
    print("Primary_Industry", Primary_Industry)
    print("Industry_category", Industry_category)

    return Company, Ticker, Primary_Industry, Industry_category

def loadDataSet2():  #加载数据集
    wb = openpyxl.load_workbook('index.xlsx')
    # 读取工作表
    ws = wb['Sheet1']
    Factor = []
    Factor.append('Operating Status')
    Factor.append('Social Factors')
    Factor.append('Political Factors')
    Factor.append('Transaction Situation')
    Operating_Status = []
    Social_Factors = []
    Political_Factors= []
    Transaction_Situation = []
    for col in ws['A']:
        Operating_Status.append(col.value)
    Operating_Status.remove('Operating Status')
    for col in ws['B']:
        Social_Factors.append(col.value)
    Social_Factors.remove('Social Factors')
    for col in ws['C']:
        Political_Factors.append(col.value)
    Political_Factors.remove('Political Factors')
    for col in ws['D']:
        Transaction_Situation.append(col.value)
    Transaction_Situation.remove('Transaction Situation')

    print("Factor", Factor)
    print("Operating_Status", Operating_Status)
    print("Social_Factors", Social_Factors)
    print("Political_Factors", Political_Factors)
    print("Transaction_Situation", Transaction_Situation)
    Operating_Status.remove(None)
    Operating_Status.remove(None)

    return  Factor, Operating_Status, Social_Factors, Political_Factors, Transaction_Situation

class Example(object):

    def __init__(self, uri, user, password):
        self._driver = GraphDatabase.driver(uri, auth=(user, password))

    def close(self):
        self._driver.close()
    #创建实体节点
    def create_One_Industry(cls, tx, Name):
        tx.run("CREATE (:Industry {Name: $Name})", Name=Name)

    def create_Company(cls, tx, Name):
        tx.run("CREATE (:Company {Name: $Name})", Name=Name)

    def create_Ticker(cls, tx, Ticker):
        tx.run("CREATE (:Ticker {Name: $Ticker})", Ticker = Ticker)

    def create_Industry(cls, tx, Industry):
        tx.run("CREATE (:Primary_Industry {Name: $Industry})", Industry = Industry)

    def create_Industry_Category(cls, tx, Industry_Category):
        tx.run("CREATE (:Industry_Category {Name: $Industry_Category})", Industry_Category = Industry_Category)


    #创建关系
    def Company_Ticker(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Company {Name: $node_name1}) "
               "MATCH (node2: Ticker {Name: $node_name2}) "
               "CREATE (node1)-[:issue]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Company_Industry(cls, tx, node_name1, node_name2):
        print(node_name1, node_name2)
        tx.run("MATCH (node1: Company {Name: $node_name1}) "
               "MATCH (node2: Primary_Industry {Name: $node_name2}) "
               "CREATE (node1)-[:engage]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Industry_Classify(cls, tx, node_name1, node_name2):
        print(node_name1, node_name2)
        tx.run("MATCH (node1: Primary_Industry {Name: $node_name1}) "
               "MATCH (node2: Industry_Category {Name: $node_name2}) "
               "CREATE (node1)-[:belong_to]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def One_Industry_Industry(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Industry {Name: $node_name1}) "
               "MATCH (node2: Industry_Category {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)


    #创建Driver
    #实体节点
    def create_Company_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Company, message)

    def create_Ticker_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Ticker, message)

    def create_Industry_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Industry, message)

    def create_Industry_Category_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Industry_Category, message)

    def create_One_Industry_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_One_Industry, message)

    #创建边

    def create_Company_Ticker_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Company_Ticker, message1, message2)

    def create_One_Industry_Industry_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.One_Industry_Industry, message1, message2)

    def create_Company_Industry_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Company_Industry, message1, message2)

    def create_Industry_Classify_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Industry_Classify, message1, message2)



#创建实体节点
    def create_Factor(cls, tx, Name):
        tx.run("CREATE (:Factor {Name: $Name})", Name=Name)

    def create_Operating_Status(cls, tx, Operating_Status):
        tx.run("CREATE (:Operating_Status {Name: $Operating_Status})", Operating_Status = Operating_Status)

    def create_Social_Factors(cls, tx, Social_Factors):
        tx.run("CREATE (:Social_Factors {Name: $Social_Factors})", Social_Factors = Social_Factors)

    def create_Political_Factors(cls, tx, Political_Factors):
        tx.run("CREATE (:Political_Factors {Name: $Political_Factors})", Political_Factors = Political_Factors)

    def create_Transaction_Situation(cls, tx, Transaction_Situation):
        tx.run("CREATE (:Transaction_Situation {Name: $Transaction_Situation})", Transaction_Situation = Transaction_Situation)

    def create_Financial_elements(cls, tx, Name):
        tx.run("CREATE (:Financial_elements {Name: $Name})", Name=Name)


    #创建关系
    def Operating_Status_index(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Factor {Name: $node_name1}) "
               "MATCH (node2: Operating_Status {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Social_Factors_index(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Factor {Name: $node_name1}) "
               "MATCH (node2: Social_Factors {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Political_Factors_index(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Factor {Name: $node_name1}) "
               "MATCH (node2: Political_Factors {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Transaction_Situation_index(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Factor {Name: $node_name1}) "
               "MATCH (node2: Transaction_Situation {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Company_Factor(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Company {Name: $node_name1}) "
               "MATCH (node2: Factor {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    #创建节点Driver
    def create_Factor_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Factor, message)

    def create_Operating_Status_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Operating_Status, message)

    def create_Social_Factors_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Social_Factors, message)

    def create_Political_Factors_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Political_Factors, message)

    def create_Transaction_Situation_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Transaction_Situation, message)

    # 创建关系Driver
    def Operating_Status_index_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Operating_Status_index, message1, message2)

    def Social_Factors_index_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Social_Factors_index, message1, message2)

    def Political_Factors_index_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Political_Factors_index, message1, message2)

    def Transaction_Situation_index_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Transaction_Situation_index, message1, message2)

    def Company_Factor_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Company_Factor, message1, message2)



# exam=Example("bolt://18.166.174.144:5002","neo4j","chengdu80uestc")

exam =  Example("bolt://localhost:7687","neo4j","myneo4j")
Company, Ticker, Primary_Industry, Industry_Category = loadDataSet()
Company1 = set(Company)
Ticker1 = set(Ticker)
Primary_Industry1 = set(Primary_Industry)
Industry_Category1 = set(Industry_Category)

# 导入节点
for each in Company1:
    exam.create_Company_driver(each)
for each in Ticker1:
    exam.create_Ticker_driver(each)

for each in Primary_Industry1:
    exam.create_Industry_driver(each)
for each in Industry_Category1:
    exam.create_Industry_Category_driver(each)

exam.create_One_Industry_driver('Industry')

#导入公司和股票的关系
for i in range(0, len(Company)):
    exam.create_Company_Ticker_driver(Company[i], Ticker[i])

#导入公司和行业的关系
exam.create_Company_Industry_driver(Company[0], Primary_Industry[0])
for i in range(1, len(Company)):
    if(Company[i] != Company[i-1]):
        exam.create_Company_Industry_driver(Company[i], Primary_Industry[i])

#行业分类
s = []
for i in range(0, len(Primary_Industry)):
    if(Primary_Industry[i] not in s):
        exam.create_Industry_Classify_driver(Primary_Industry[i], Industry_Category[i])
        s.append(Primary_Industry[i])

for item in Industry_Category1:
    exam.create_One_Industry_Industry_driver('Industry', item)

Factor, Operating_Status, Social_Factors, Political_Factors, Transaction_Situation = loadDataSet2()
ss = []
for each in Company1:
    ss.append(each)
for j in range(0, len(ss)):
    for i in range(0, len(Factor)):
        exam.create_Factor_driver(Factor[i]+"_"+ss[j])

    for i in range(0, len(Operating_Status)):
        exam.create_Operating_Status_driver(Operating_Status[i]+"_"+ss[j])
    for i in range(0, len(Social_Factors)):
        exam.create_Social_Factors_driver(Social_Factors[i]+"_"+ss[j])
    for i in range(0, len(Political_Factors)):
        exam.create_Political_Factors_driver(Political_Factors[i]+"_"+ss[j])
    for i in range(0, len(Transaction_Situation)):
        exam.create_Transaction_Situation_driver(Transaction_Situation[i]+"_"+ss[j])

    for i in range(0, len(Factor)):
        exam.Company_Factor_driver(ss[j], Factor[i]+"_"+ss[j])

    for i in range(0, len(Operating_Status)):
        exam.Operating_Status_index_driver(Factor[0]+"_"+ss[j], Operating_Status[i]+"_"+ss[j])

    for i in range(0, len(Social_Factors)):
        exam.Social_Factors_index_driver(Factor[1]+"_"+ss[j], Social_Factors[i]+"_"+ss[j])

    for i in range(0, len(Political_Factors)):
        exam.Political_Factors_index_driver(Factor[2]+"_"+ss[j], Political_Factors[i]+"_"+ss[j])

    for i in range(0, len(Transaction_Situation)):
        exam.Transaction_Situation_index_driver(Factor[3]+"_"+ss[j], Transaction_Situation[i]+"_"+ss[j])

