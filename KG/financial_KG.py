import numpy
import pandas
import openpyxl
from neo4j import GraphDatabase


def loadDataSet():  #加载数据集
    wb = openpyxl.load_workbook('Company_relationship.xlsx')
    # 读取工作表
    ws = wb['Sheet1']
    Company = []
    Ticker = []
    Primary_Industry = []
    Industry_category = []
    for col in ws['A']:
        Company.append(col.value)
    Company.remove('Company Name')
    for col in ws['B']:
        Ticker.append(col.value)
    Ticker.remove('Ticker')
    for col in ws['C']:
        Primary_Industry.append(col.value)
    Primary_Industry.remove('Primary Industry')
    for col in ws['D']:
        Industry_category.append(col.value)
    Industry_category.remove('Industry category')


    print("Company", Company)
    print("Ticker", Ticker)
    print("Primary_Industry", Primary_Industry)
    print("Industry_category", Industry_category)

    return Company, Ticker, Primary_Industry, Industry_category


class Example(object):

    def __init__(self, uri, user, password):
        self._driver = GraphDatabase.driver(uri, auth=(user, password))

    def close(self):
        self._driver.close()
    #创建实体节点
    def create_Company(cls, tx, Name):
        tx.run("CREATE (:Company {Name: $Name})", Name=Name)

    def create_Ticker(cls, tx, Ticker):
        tx.run("CREATE (:Ticker {Name: $Ticker})", Ticker = Ticker)

    def create_Industry(cls, tx, Industry):
        tx.run("CREATE (:Industry {Name: $Industry})", Industry = Industry)

    def create_Industry_Category(cls, tx, Industry_Category):
        tx.run("CREATE (:Industry_Category {Name: $Industry_Category})", Industry_Category = Industry_Category)




    #创建关系
    def Company_Ticker(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Company {Name: $node_name1}) "
               "MATCH (node2: Ticker {Name: $node_name2}) "
               "CREATE (node1)-[:issue]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Company_Industry(cls, tx, node_name1, node_name2):
        print(node_name1, node_name2)
        tx.run("MATCH (node1: Company {Name: $node_name1}) "
               "MATCH (node2: Industry {Name: $node_name2}) "
               "CREATE (node1)-[:engage]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Industry_Classify(cls, tx, node_name1, node_name2):
        print(node_name1, node_name2)
        tx.run("MATCH (node1: Industry {Name: $node_name1}) "
               "MATCH (node2: Industry_Category {Name: $node_name2}) "
               "CREATE (node1)-[:belong_to]->(node2)",
               node_name1=node_name1, node_name2=node_name2)


    #创建Driver
    def create_Company_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Company, message)

    def create_Ticker_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Ticker, message)

    def create_Industry_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Industry, message)

    def create_Industry_Category_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Industry_Category, message)

    def create_Company_Ticker_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Company_Ticker, message1, message2)

    def create_Company_Industry_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Company_Industry, message1, message2)

    def create_Industry_Classify_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Industry_Classify, message1, message2)

exam=Example("bolt://localhost:7687","neo4j","myneo4j")

Company, Ticker, Primary_Industry, Industry_Category = loadDataSet()
Company1 = set(Company)
Ticker1 = set(Ticker)
Primary_Industry1 = set(Primary_Industry)
Industry_Category1 = set(Industry_Category)

#导入节点
for each in Company1:
    exam.create_Company_driver(each)
for each in Ticker1:
    exam.create_Ticker_driver(each)
for each in Primary_Industry1:
    exam.create_Industry_driver(each)
for each in Industry_Category1:
    exam.create_Industry_Category_driver(each)

#导入公司和股票的关系
for i in range(0, len(Company)):
    exam.create_Company_Ticker_driver(Company[i], Ticker[i])

#导入公司和行业的关系
exam.create_Company_Industry_driver(Company[0], Primary_Industry[0])
for i in range(1, len(Company)):
    if(Company[i] != Company[i-1]):
        exam.create_Company_Industry_driver(Company[i], Primary_Industry[i])

#行业分类
s = []
for i in range(0, len(Primary_Industry)):
    if(Primary_Industry[i] not in s):
        exam.create_Industry_Classify_driver(Primary_Industry[i], Industry_Category[i])
        s.append(Primary_Industry[i])


