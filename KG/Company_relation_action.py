import numpy
import pandas
import openpyxl
from neo4j import GraphDatabase


def loadDataSet():  #加载数据集
    wb = openpyxl.load_workbook('Company_relationship.xlsx')
    # 读取工作表
    ws = wb['Sheet1']
    Company = []
    Ticker = []
    Primary_Industry = []
    Industry_category = []
    for col in ws['A']:
        Company.append(col.value)
    Company.remove('Company Name')
    for col in ws['B']:
        Ticker.append(col.value)
    Ticker.remove('Ticker')
    for col in ws['C']:
        Primary_Industry.append(col.value)
    Primary_Industry.remove('Primary Industry')
    for col in ws['D']:
        Industry_category.append(col.value)
    Industry_category.remove('Industry category')


    print("Company", Company)
    print("Ticker", Ticker)
    print("Primary_Industry", Primary_Industry)
    print("Industry_category", Industry_category)

    return Company, Ticker, Primary_Industry, Industry_category

def loadDataSet2():  # 函数功能为打开文本文件并逐行读取
    company = []
    company2 = []
    relation = []
    effect = []
    comp = open('company_names.txt')
    for line in comp.readlines():
        company.append(line.strip('\n'))

    fr = open('triple_with_comp_10.txt', 'r', encoding='UTF-8')
    x = []
    for line in fr.readlines():
        x.append(line.strip('\n'))
    x = set(x)
    print(x)
    for each in x:
        curLine = each.split('|')

        company2.append(curLine[0])
        relation.append(curLine[1])
        effect.append(curLine[2].strip('\n'))

    return company, company2, effect, relation


class Example(object):

    def __init__(self, uri, user, password):
        self._driver = GraphDatabase.driver(uri, auth=(user, password))

    def close(self):
        self._driver.close()
    #创建实体节点
    def create_One_Industry(cls, tx, Name):
        tx.run("CREATE (:Industry {Name: $Name})", Name=Name)

    def create_Company(cls, tx, Name):
        tx.run("CREATE (:Company {Name: $Name})", Name=Name)

    def create_Ticker(cls, tx, Ticker):
        tx.run("CREATE (:Ticker {Name: $Ticker})", Ticker = Ticker)

    def create_Industry(cls, tx, Industry):
        tx.run("CREATE (:Primary_Industry {Name: $Industry})", Industry = Industry)

    def create_Industry_Category(cls, tx, Industry_Category):
        tx.run("CREATE (:Industry_Category {Name: $Industry_Category})", Industry_Category = Industry_Category)


    #创建关系
    def Company_Ticker(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Company {Name: $node_name1}) "
               "MATCH (node2: Ticker {Name: $node_name2}) "
               "CREATE (node1)-[:issue]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Company_Industry(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Company {Name: $node_name1}) "
               "MATCH (node2: Primary_Industry {Name: $node_name2}) "
               "CREATE (node1)-[:engage]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Industry_Classify(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Primary_Industry {Name: $node_name1}) "
               "MATCH (node2: Industry_Category {Name: $node_name2}) "
               "CREATE (node1)-[:belong_to]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def One_Industry_Industry(cls, tx, node_name1, node_name2):
        tx.run("MATCH (node1: Industry {Name: $node_name1}) "
               "MATCH (node2: Industry_Category {Name: $node_name2}) "
               "CREATE (node1)-[:include]->(node2)",
               node_name1=node_name1, node_name2=node_name2)


    #创建Driver
    #实体节点
    def create_Company_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Company, message)

    def create_Ticker_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Ticker, message)

    def create_Industry_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Industry, message)

    def create_Industry_Category_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Industry_Category, message)

    def create_One_Industry_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_One_Industry, message)

    #创建边

    def create_Company_Ticker_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Company_Ticker, message1, message2)

    def create_One_Industry_Industry_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.One_Industry_Industry, message1, message2)

    def create_Company_Industry_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Company_Industry, message1, message2)

    def create_Industry_Classify_driver(self, message1, message2):
        with self._driver.session() as session:
            session.write_transaction(self.Industry_Classify, message1, message2)

    ###########################Copmany_action##############################################

    def create_Effect(cls, tx, Effect):
        tx.run("CREATE (:Effect {Name: $Effect})", Effect = Effect)

    def Company_Effect(cls, tx, node_name1, node_name2, relation):
        tx.run("MATCH (node1: Company {Name: $node_name1}) "
               "MATCH (node2: Effect {Name: $node_name2}) "
               "CREATE (node1)-[: " +relation+"]->(node2)",
               node_name1=node_name1, node_name2=node_name2)

    def Company_Company(cls, tx, node_name1, node_name2, relation):
        tx.run("MATCH (node1: Company {Name: $node_name1}) "
               "MATCH (node2: Company {Name: $node_name2}) "
               "CREATE (node1)-[: "+relation+"]->(node2)",
               node_name1=node_name1, node_name2=node_name2)


    def create_Effect_driver(self, message):
        with self._driver.session() as session:
            session.write_transaction(self.create_Effect, message)

    def Company_Company_driver(self, message1, message2, message3):
        with self._driver.session() as session:
            session.write_transaction(self.Company_Company, message1, message2, message3)

    def Company_Effect_driver(self, message1, message2, message3):
        with self._driver.session() as session:
            session.write_transaction(self.Company_Effect, message1, message2, message3)


exam=Example("bolt://localhost:7687","neo4j","myneo4j")

Company, Ticker, Primary_Industry, Industry_Category = loadDataSet()

company, company2, effect, realtion = loadDataSet2()

Company1 = set(Company)
Ticker1 = set(Ticker)
Primary_Industry1 = set(Primary_Industry)
Industry_Category1 = set(Industry_Category)

# 导入节点
for each in Company1:
    exam.create_Company_driver(each)
for each in Ticker1:
    exam.create_Ticker_driver(each)

for each in Primary_Industry1:
    exam.create_Industry_driver(each)
for each in Industry_Category1:
    exam.create_Industry_Category_driver(each)

exam.create_One_Industry_driver('Industry')

#导入公司和股票的关系
for i in range(0, len(Company)):
    exam.create_Company_Ticker_driver(Company[i], Ticker[i])

#导入公司和行业的关系
exam.create_Company_Industry_driver(Company[0], Primary_Industry[0])
for i in range(1, len(Company)):
    if(Company[i] != Company[i-1]):
        exam.create_Company_Industry_driver(Company[i], Primary_Industry[i])

#行业分类
s = []
for i in range(0, len(Primary_Industry)):
    if(Primary_Industry[i] not in s):
        exam.create_Industry_Classify_driver(Primary_Industry[i], Industry_Category[i])
        s.append(Primary_Industry[i])

for item in Industry_Category1:
    exam.create_One_Industry_Industry_driver('Industry', item)

j = 0
for each in realtion:
    if ' ' in each:
        for i in range(0,len(each)):
            if each[i] == ' ':
                t = list(each)
                t[i] = '_'
                each = ''.join(t)
        for i in range(0, len(each)):
            if each[i] == '-':
                t = list(each)
                t[i] = '_'
                each = ''.join(t)
                print(each)

    realtion[j] = each
    j = j + 1

names = set()


j = 0
xx = 0
for i in range(0, len(company2)):
    for j in range(0, len(Company)):
        if company2[i] in Company[j]:
            company2[i] = Company[j]
            xx = xx + 1
            break
print("XX", xx)
for i in range(0, len(effect)):
    for j in range(0, len(Company)):
        if effect[i] in Company[j]:
            effect[i] = Company[j]
            break

for i in range(0, len(company2)):
    if company2[i] in Company:
        if effect[i] != company2[i]:
            if effect[i] in Company:
                exam.Company_Company_driver(company2[i], effect[i], realtion[i])
            else:
                print("aaaaaaaaaaaaaaaaaaaaaaaaa")
                exam.create_Effect_driver(effect[i])
                exam.Company_Effect_driver(company2[i], effect[i], realtion[i])

f = open('file.txt', 'r')
obj = json.load(f)


