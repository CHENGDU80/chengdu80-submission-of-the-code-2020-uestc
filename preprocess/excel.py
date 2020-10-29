from openpyxl import Workbook, load_workbook


wb = load_workbook('company_info.xlsx')
table = wb['industrial_relation']

companies = []
cat = set()
for r in range(2, table.max_row+1):
    company = {}
    company['name'] = table.cell(r, 1).value
    company['ticker'] = table.cell(r, 2).value
    company['industry'] = table.cell(r, 3).value
    company['category'] = table.cell(r, 4).value
    cat.add(company['category'])
    companies.append(company)
    print(company)

with open('companies.json', 'w') as f:
    import json
    json.dump(companies, f, indent=2)
